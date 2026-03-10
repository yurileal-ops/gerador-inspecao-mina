import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, NamedStyle
from openpyxl.utils import get_column_letter
import os
import sys
import subprocess
import threading

def processar_relatorio():
    """Função que processa o relatório TCLD DA 2"""
    # Abrir diálogo para selecionar arquivo
    root = tk.Tk()
    root.withdraw()  # Esconde a janela principal
    root.attributes('-topmost', True)  # Coloca sempre na frente
    
    arquivo_entrada = filedialog.askopenfilename(
        title="Selecione o arquivo RELATORIO.xlsx (TCLD DA 2)",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        initialdir=os.path.expanduser("~\\Desktop")
    )
    
    if not arquivo_entrada:
        messagebox.showwarning("Cancelado", "Nenhum arquivo foi selecionado.")
        root.destroy()
        return
    
    # Validar se arquivo existe
    if not os.path.exists(arquivo_entrada):
        messagebox.showerror("Erro", f"Arquivo não encontrado: {arquivo_entrada}")
        root.destroy()
        return
    
    arquivo_saida = os.path.join(os.path.dirname(arquivo_entrada), 'INSPECAO_SEMANAL_TCLD_DA_2.xlsx')
    
    try:
        # 1. Ler o arquivo Excel usando pandas
        df = pd.read_excel(arquivo_entrada)
        root.update()  # Atualizar mensagem
        
        # Verificar se as colunas necessárias existem
        colunas_necessarias = ['Descrição', 'TextPrioridade', 'Texto do item', 'Nota', 'Data de criação']
        for col in colunas_necessarias:
            if col not in df.columns:
                root.destroy()
                messagebox.showerror("Erro", f"Coluna '{col}' não encontrada no arquivo Excel.\n\nColunas necessárias:\n{', '.join(colunas_necessarias)}")
                return
        
        # 8. Ignorar linhas vazias ou dados inválidos
        df = df.dropna(subset=['Descrição', 'Texto do item'])
        df = df[df['Descrição'].str.strip() != '']
        df = df[df['Texto do item'].str.strip() != '']
        
        # 2. Extrair automaticamente a CORREIA com regex
        df['Correia'] = df['Descrição'].str.extract(r'([\dA-Z]+(?:CV|AL)\d+)', expand=False).str.strip()
        
        # Definir os sistemas TCLD DA 2 e suas correias
        sistemas = {
            'BRITAGEM - Transportadores': ['03CV001', '03CV002', '03CV003', '03CV004', '03CV005', '03CV006', '05CV001', '05CV002'],
            'BRITAGEM - ALS': ['03AL001', '03AL002', '03AL003', '03AL004', '03AL005', '03AL006', '03AL007', '03AL008', '03AL009', '03AL010', '03AL011', '03AL012', '03AL013'],
            'USINA - ALS': ['05AL001', '05AL002'],
            'USINA - Transportadores': ['05CV003', '05CV004', '05CV005', '05CV006'],
            'TCLD SUL 1 - Transportadores': ['02CV003', '02CV004', '02CV005'],
            'FILTRAGEM PDER SUL': ['52CV081', '52CV082', '52CV083', '52CV084', '52CV085', '52CV086'],
            'DESCARACTERIZAÇÃO DE BARRAGEM': ['49CV115', '49CV116', '53CV088', '53CV089', '53CV090', '53CV097']
        }
        
        # Criar dicionário correia -> sistema
        correia_to_sistema = {}
        for sistema, correias in sistemas.items():
            for correia in correias:
                correia_to_sistema[correia] = sistema
        
        # Filtrar apenas correias válidas
        df = df[df['Correia'].isin(correia_to_sistema.keys())]
        
        # Adicionar coluna Sistema
        df['Sistema'] = df['Correia'].map(correia_to_sistema)
        
        # 3. Extrair o número do CAVALETE
        df['Cavalete'] = df['Texto do item'].str.split('-').str[0].str.strip()
        df['Cavalete'] = pd.to_numeric(df['Cavalete'], errors='coerce')
        df = df.dropna(subset=['Cavalete'])
        df['Cavalete'] = df['Cavalete'].astype(int)
        
        # Extrair o tipo do rolo
        df['Tipo'] = df['Texto do item'].str.split('-').str[1].str.strip()
        
        # 4. Ordenar os dados
        df = df.sort_values(by=['Sistema', 'Correia', 'Cavalete'])
        
        # 5. Agrupar os itens
        grouped = df.groupby(['Sistema', 'Correia'])
        
        # 9. e 10. Gerar novo arquivo Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Inspeção Semanal TCLD DA 2"
        
        center_alignment = Alignment(horizontal='center', vertical='center')
        date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
        
        # Cabeçalhos
        ws['A1'] = 'PRIORIDADE'
        ws['B1'] = 'TEXTO DO ITEM'
        ws['C1'] = 'CORREIA'
        ws['O1'] = 'Nº NOTA'
        ws['P1'] = 'DATA INSPEÇÃO'
        
        for col in ['A', 'B', 'C', 'O', 'P']:
            ws[f'{col}1'].alignment = center_alignment
        
        current_row = 2
        
        for (sistema, correia), group in grouped:
            for _, row in group.iterrows():
                ws.cell(row=current_row, column=1).value = row['TextPrioridade']
                ws.cell(row=current_row, column=1).alignment = center_alignment
                
                ws.cell(row=current_row, column=2).value = row['Texto do item']
                ws.cell(row=current_row, column=2).alignment = center_alignment
                
                ws.cell(row=current_row, column=3).value = row['Correia']
                ws.cell(row=current_row, column=3).alignment = center_alignment
                
                ws.cell(row=current_row, column=4).value = f'=LEFT(B{current_row},FIND("-",B{current_row})-1)'
                ws.cell(row=current_row, column=4).alignment = center_alignment
                
                ws.cell(row=current_row, column=5).value = f'=MID(B{current_row},FIND("-",B{current_row})+1,FIND("-",B{current_row},FIND("-",B{current_row})+1)-FIND("-",B{current_row})-1)'
                ws.cell(row=current_row, column=5).alignment = center_alignment
                
                ws.cell(row=current_row, column=7).value = f'=IF(ISERROR(FIND("-D-",B{current_row})),"","X")'
                ws.cell(row=current_row, column=7).alignment = center_alignment
                
                ws.cell(row=current_row, column=8).value = f'=IF(ISERROR(FIND("-C-",B{current_row})),"","X")'
                ws.cell(row=current_row, column=8).alignment = center_alignment
                
                ws.cell(row=current_row, column=9).value = f'=IF(ISERROR(FIND("-E-",B{current_row})),"","X")'
                ws.cell(row=current_row, column=9).alignment = center_alignment
                
                ws.cell(row=current_row, column=11).value = f'=IF(ISERROR(FIND("P1",A{current_row})),"","X")'
                ws.cell(row=current_row, column=11).alignment = center_alignment
                
                ws.cell(row=current_row, column=12).value = f'=IF(ISERROR(FIND("P2",A{current_row})),"","X")'
                ws.cell(row=current_row, column=12).alignment = center_alignment
                
                ws.cell(row=current_row, column=13).value = f'=IF(ISERROR(FIND("P3",A{current_row})),"","X")'
                ws.cell(row=current_row, column=13).alignment = center_alignment
                
                dano_formula = f'=IF(ISERROR(FIND("RO",B{current_row})),IF(ISERROR(FIND("CD",B{current_row})),IF(ISERROR(FIND("CG",B{current_row})),IF(ISERROR(FIND("AG",B{current_row})),IF(ISERROR(FIND("RT",B{current_row})),IF(ISERROR(FIND("RF",B{current_row})),IF(ISERROR(FIND("AM",B{current_row})),IF(ISERROR(FIND("RP",B{current_row})),"","ROLO FORA DE POSIÇÃO"),"ACUMULO DE MATERIAL"),"ROLO FALTANTE"),"ROLETE TRAVADO"),"DESGASTE NATURAL"),"CILINDRO GASTO"),"DEGOLADO"),"ROLAMENTO DANIFICADO")'
                ws.cell(row=current_row, column=14).value = dano_formula
                ws.cell(row=current_row, column=14).alignment = center_alignment
                
                ws.cell(row=current_row, column=15).value = row['Nota']
                ws.cell(row=current_row, column=15).alignment = center_alignment
                
                ws.cell(row=current_row, column=16).value = row['Data de criação']
                ws.cell(row=current_row, column=16).alignment = center_alignment
                ws.cell(row=current_row, column=16).style = date_style
                
                current_row += 1
            current_row += 1
        
        # Ajustar largura das colunas
        for col_num in [1, 2, 3, 15, 16]:
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].auto_size = True
        
        # Salvar o arquivo
        wb.save(arquivo_saida)
        root.destroy()
        
        # Mostrar mensagem de sucesso e opção de abrir
        response = messagebox.askyesno(
            "Sucesso", 
            f"Arquivo '{os.path.basename(arquivo_saida)}' gerado com sucesso!\n\nDeseja abrir o arquivo agora?"
        )
        
        if response:
            # Abrir arquivo com Excel ou aplicativo padrão
            try:
                os.startfile(arquivo_saida)
            except Exception as e:
                messagebox.showerror("Erro", f"Não foi possível abrir o arquivo: {str(e)}")
        
    except Exception as e:
        root.destroy()
        messagebox.showerror("Erro ao processar", f"Erro ao processar: {str(e)}\n\nVerifique se o arquivo está em formato correto.")

if __name__ == "__main__":
    processar_relatorio()
