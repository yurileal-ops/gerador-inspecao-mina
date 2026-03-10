import pandas as pd
import openpyxl

# Caminho do arquivo de entrada
arquivo_entrada = 'relatorio.xlsx'

# Caminho do arquivo de saída
arquivo_saida = 'INSPECAO_SEMANAL.xlsx'

# 1. Ler o arquivo Excel usando pandas
try:
    df = pd.read_excel(arquivo_entrada)
    print("Arquivo Excel lido com sucesso.")
except FileNotFoundError:
    print(f"Erro: Arquivo '{arquivo_entrada}' não encontrado.")
    exit(1)
except Exception as e:
    print(f"Erro ao ler o arquivo Excel: {e}")
    exit(1)

# Verificar se as colunas necessárias existem
colunas_necessarias = ['Descrição', 'TextPrioridade', 'Texto do item', 'Nota', 'Data de criação']
for col in colunas_necessarias:
    if col not in df.columns:
        print(f"Erro: Coluna '{col}' não encontrada no arquivo Excel.")
        exit(1)

# 8. Ignorar linhas vazias ou dados inválidos
# Remover linhas onde 'Descrição' ou 'Texto do item' estão vazios ou NaN
df = df.dropna(subset=['Descrição', 'Texto do item'])
df = df[df['Descrição'].str.strip() != '']
df = df[df['Texto do item'].str.strip() != '']

# 2. Extrair automaticamente a CORREIA a partir da coluna "Descrição"
# Assumindo que a correia é a parte antes do primeiro '-'
df['Correia'] = df['Descrição'].str.split('-').str[0].str.strip()

# Definir os sistemas e suas correias
sistemas = {
    'FAZENDÃO': ['11CV56', '11LK02', '11LL01', '11LL03'],
    'ALEGRIA SUL 02': ['11CV68', '11CV67', '11CR05', '02CV09', '11CR19', '11LL07', '11LL09', '11CR14', '11CV07', '11CV69', '11LK03'],
    'ALEGRIA CENTRO': ['11CV21', '02CV37', '02CV38', '02CV39', '02CV40', '02CV41', '11CR13'],
    'ALEGRIA CENTRO 64': ['11CV20', '11CV64', '11LK01', '11LL04', '11CV60', '11LL06'],
    'ALEGRIA 345': ['11CV23', '11CV24', '11CR10'],
    'ALEGRIA SUL 01': ['11CV72', '02CV02', '11CR12']
}

# Criar dicionário correia -> sistema
correia_to_sistema = {}
for sistema, correias in sistemas.items():
    for correia in correias:
        correia_to_sistema[correia] = sistema

# Filtrar apenas correias válidas
df = df[df['Correia'].isin(correia_to_sistema.keys())]

# Adicionar coluna Sistema
df['Sistema'] = df['Correia'].map(correia_to_sistema)

# 3. Extrair o número do CAVALETE a partir da coluna "Texto do item"
# Assumindo que o cavalete é o número antes do primeiro '-'
df['Cavalete'] = df['Texto do item'].str.split('-').str[0].str.strip()
# Converter para int, ignorando erros (caso não seja número)
df['Cavalete'] = pd.to_numeric(df['Cavalete'], errors='coerce')
# Remover linhas onde Cavalete não é um número válido
df = df.dropna(subset=['Cavalete'])
df['Cavalete'] = df['Cavalete'].astype(int)

# Extrair o tipo do rolo (GR, RC, RI, RR, etc.) para contagem
df['Tipo'] = df['Texto do item'].str.split('-').str[1].str.strip()

# 4. Ordenar os dados primeiro pelo sistema, depois pela correia e depois pelo número do cavalete
df = df.sort_values(by=['Sistema', 'Correia', 'Cavalete'])

# 5. Agrupar os itens por sistema e correia
grouped = df.groupby(['Sistema', 'Correia'])

# 9. e 10. Gerar um novo arquivo Excel chamado INSPECAO_SEMANAL.xlsx
# Criar um novo workbook sem modelo, organizando os dados por sistema e correia
from openpyxl import Workbook
from openpyxl.styles import Alignment, NamedStyle
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Inspeção Semanal"

# Definir estilo centralizado
center_alignment = Alignment(horizontal='center', vertical='center')

# Definir estilo para data
date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')

# Cabeçalhos
ws['A1'] = 'PRIORIDADE'
ws['B1'] = 'TEXTO DO ITEM'
ws['C1'] = 'CORREIA'
ws['O1'] = 'Nº NOTA'
ws['P1'] = 'DATA INSPEÇÃO'

# Aplicar estilos aos cabeçalhos
for col in ['A', 'B', 'C', 'O', 'P']:
    ws[f'{col}1'].alignment = center_alignment

current_row = 2  # Começar na linha 2

# percorre grupos e escreve
for (sistema, correia), group in grouped:
    for _, row in group.iterrows():
        ws.cell(row=current_row, column=1).value = row['TextPrioridade']
        ws.cell(row=current_row, column=1).alignment = center_alignment

        ws.cell(row=current_row, column=2).value = row['Texto do item']
        ws.cell(row=current_row, column=2).alignment = center_alignment

        ws.cell(row=current_row, column=3).value = row['Correia']
        ws.cell(row=current_row, column=3).alignment = center_alignment

        ws.cell(row=current_row, column=15).value = row['Nota']
        ws.cell(row=current_row, column=15).alignment = center_alignment

        ws.cell(row=current_row, column=16).value = row['Data de criação']
        ws.cell(row=current_row, column=16).alignment = center_alignment
        ws.cell(row=current_row, column=16).style = date_style

        current_row += 1
    # pular uma linha em branco para separar correias
    current_row += 1

# Ajustar largura das colunas
for col_num in [1, 2, 3, 15, 16]:
    column_letter = get_column_letter(col_num)
    ws.column_dimensions[column_letter].auto_size = True

# Salvar o arquivo
try:
    wb.save(arquivo_saida)
    print(f"Arquivo Excel de saída '{arquivo_saida}' gerado com sucesso.")
except Exception as e:
    print(f"Erro ao gerar o arquivo Excel de saída: {e}")

