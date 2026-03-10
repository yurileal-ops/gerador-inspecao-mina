import openpyxl

wb = openpyxl.load_workbook('MODELO_INSPECAO.xlsx', data_only=False)
ws = wb.active
print('max', ws.max_row, ws.max_column)
for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
    print(row)
